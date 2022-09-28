from glob import glob
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from pyfiglet import Figlet
from selenium.webdriver.common.by import By
from openpyxl import Workbook
import time

from bs4 import BeautifulSoup



f = Figlet(font='big')
print(f.renderText('Google Map Scrapper'))
searchText = ""
while searchText == "":
    searchText = input("Please enter your search : ") 

driver = webdriver.Chrome('./chromedriver')

driver.get(f"https://www.google.com/search?q={searchText}&tbm=lcl")
#1 Step
divList = driver.find_elements(By.CLASS_NAME, "rllt__details")


wb = Workbook()
ws = wb.active

index = 0
limit = -1

def check_all_item(divList):
    global index
    for div in divList:
        div.click()
        time.sleep(2)
        index += 1
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        boxSoup = soup.find(name='div', class_="xpdopen")
        activeSoup = soup.find(class_="rllt__local-item-selected")
        # activeSoup = BeautifulSoup(activeSoup, 'html.parser')
        nameSoup = boxSoup.find(name="h2", class_="qrShPb")
        name = nameSoup.text if nameSoup is not None else 'n/a'
        directionSoup = soup.find(class_='ab_button',string="Directions")
        direction = directionSoup.get('data-url') if directionSoup is not None else '-'
        if direction != '-':
            direction = "https://www.google.com/" + direction
        
        websiteSoup = soup.find(class_='ab_button',string="Website")
        website = websiteSoup.get('href') if websiteSoup is not None else '-'

        ratingSoup = activeSoup.find(class_='YDIN4c YrbPuc')
        rating = ratingSoup.text if ratingSoup is not None else '-'

        ratingOutOfSoup = activeSoup.find(class_='HypWnf YrbPuc')
        ratingOutOf = ratingOutOfSoup.text if ratingOutOfSoup is not None else '-'
        # print(soup.find(_class='YDIN4c YrbPuc'))
        # if index>0:
        #     driver.close()
        #     break

        descrSoup = soup.find(name="div",class_='kno-rdesc')
        descr = descrSoup.text if descrSoup is not None else '-'

        addressSoup = boxSoup.find(name="a",string="Address")
        address = addressSoup.parent.parent.text if addressSoup is not None else '-'

        phoneSoup = boxSoup.find(name="a",string="Phone")
        phone = phoneSoup.parent.parent.text if phoneSoup is not None else '-'
        if phone is not None:
            phone = phone.replace('Phone: ', '')

    
        el = soup.find_all(lambda tag:tag is not None and tag.name=="div" and ": " in tag.text)
        ep = []
        for e in el:
            if e.get('class') is  None:
                continue
            if "wDYxhc" in e.get('class'):
                ep.append(e.text)
                # print(e)
        meta = "\n".join(ep)

        ws.cell(index,1,str(name))
        ws.cell(index,2,str(rating))
        ws.cell(index,3,str(ratingOutOf))
        ws.cell(index,4,str(website))
        ws.cell(index,5,str(phone))
        ws.cell(index,6,str(address))
        ws.cell(index,7,'=HYPERLINK("{}", "{}")'.format(str(direction), "Get Direction"))
        ws.cell(index,8,str(descr))
        ws.cell(index,9,str(meta))
        print(f"{index} => {name}")


check_all_item(divList)

# 2 Step
nextButton = driver.find_element(By.ID,'pnnext')

while nextButton:
    nextButton.click()
    time.sleep(3)
    divList = driver.find_elements(By.CLASS_NAME, "rllt__details")
    check_all_item(divList)
    nextButton = driver.find_element(By.ID,'pnnext')



wb.save(f"{searchText}.xlsx")


time.sleep(3)

driver.close()


# search_bar.clear()
# search_bar.send_keys("getting started with python")
# search_bar.send_keys(Keys.RETURN)